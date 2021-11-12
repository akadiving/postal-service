from datetime import datetime
from rest_framework import generics
import requests
from rest_framework import permissions, filters, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.generics import ListAPIView, CreateAPIView, RetrieveAPIView, \
    RetrieveUpdateAPIView, RetrieveDestroyAPIView
from .utils import generate_zpl
from rest_framework.decorators import api_view, permission_classes
from .models import Item, Manifest
from .serializers import ItemSerializer, ManifestSerializer
from rest_framework.response import Response
from django.http import HttpResponse
from django.http import JsonResponse
from django.db.models import Count
import xlwt
from openpyxl import load_workbook
from base64 import b64decode
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.views.decorators.clickjacking import xframe_options_exempt
User = get_user_model()
# Create your views here.


class IsOwnerOrReadOnly(permissions.BasePermission):
    """
    Object-level permission to only allow owners of an object to edit it.
    Assumes the model instance has an `owner` attribute.
    """

    def has_object_permission(self, request, view, obj):
        # Read permissions are allowed to any request,
        # so we'll always allow GET, HEAD or OPTIONS requests.
        # if request.method in permissions.SAFE_METHODS:
        # return True
        if request.user.is_superuser:
            return True

        # Instance must have an attribute named `owner`.
        return obj.owner == request.user

    def has_permission(self, request, view):
        if request.user.groups.filter(name='Company').exists() or request.user.is_superuser:
            return True
        return False

class IsOwnerFilterBackend(filters.BaseFilterBackend):
    """
    Filter that only allows users to see their own objects.
    """

    def filter_queryset(self, request, queryset, view):
        return queryset.filter(company=request.user.company_name)

class IteamSearchView(ListAPIView, IsOwnerOrReadOnly):
    permission_classes = [IsOwnerOrReadOnly, IsAuthenticated]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['^barcode', '^owner__username',
                     '^sender_name', '^receiver_id',
                     '^sender_surname', 
                     '^receiver_name', '^receiver_surname',
                     '^manifest_number__manifest_code'
                     ]


    def get_queryset(self):
        owner = self.request.user
        if owner.is_anonymous:
            return None
        return self.queryset.filter(company=owner.company_name)

class IteamSearchDeliveredView(ListAPIView, IsOwnerOrReadOnly):
    permission_classes = [IsOwnerOrReadOnly, IsAuthenticated]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['^barcode', '^owner__username',
                     '^sender_name', '^receiver_id',
                     '^sender_surname', 
                     '^receiver_name', '^receiver_surname',
                     '^manifest_number__manifest_code'
                     ]
    pagination_class = None

    def get_queryset(self):
        owner = self.request.user
        if owner.is_anonymous:
            return None
        return self.queryset.filter(delivered=True, company=owner.company_name)

class ItemListView(ListAPIView, IsOwnerOrReadOnly):
    permission_classes = [IsOwnerOrReadOnly, IsAuthenticated]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    pagination_class = None

    def get_queryset(self):
        owner = self.request.user
        if owner.is_anonymous:
            return None
        return self.queryset.filter(company=owner.company_name)

@api_view(['GET', 'POST', 'DELETE'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def stickers(request):
    # Check if user is authenticated.
    if request.user.is_anonymous:
        message = {'message': 'You are not allowed to visit this page'}
        return JsonResponse(message, safe=False)
    elif request.method != 'POST':
        message = {'message': 'You can use only POST method'}
        return JsonResponse(message, safe=False)
    obj = Item.objects.filter(id__in=request.data['id']).order_by('created_at')
    serializer = ItemSerializer(data=obj, many=True, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET', 'POST'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def filter_by_date(request):
    if request.user.is_anonymous:
        message = {'message': 'You are not allowed to visit this page'}
        return JsonResponse(message, safe=False)
    
    owner = request.user

    if request.user.is_superuser and request.data['start'] and not request.data['end']:
        item_list = Item.objects.filter(created_at__gte=request.data['start'], company=owner.company_name)

    elif request.user.is_superuser and request.data['end'] and not request.data['start']:
        item_list = Item.objects.filter(created_at__lte=request.data['end'], company=owner.company_name)

    elif request.user.is_superuser and request.data['end'] and request.data['start']:
        item_list = Item.objects.filter(created_at__range=[request.data['start'], request.data['end']], company=owner.company_name)

    elif request.user.groups.filter(name='Company').exists() and request.data['start'] and not request.data['end']:
        item_list = Item.objects.filter(owner=request.user, created_at__gte=request.data['start'], company=owner.company_name)

    elif request.user.groups.filter(name='Company').exists() and request.data['end'] and not request.data['start']:
        item_list = Item.objects.filter(owner=request.user, created_at__lte=request.data['end'], company=owner.company_name)

    elif request.user.groups.filter(name='Company').exists() and request.data['end'] and request.data['start']:
        item_list = Item.objects.filter(owner=request.user, created_at__range=[request.data['start'], request.data['end']], company=owner.company_name)
    else:
        return None
        
    serializer = ItemSerializer(data=item_list, many=True)

    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.data, status=status.HTTP_200_OK)

class AddItemView(CreateAPIView, IsOwnerOrReadOnly):
    permission_classes = [IsAuthenticated]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user,
                        company=self.request.user.company_name)

class UpdateItemView(RetrieveUpdateAPIView, IsOwnerOrReadOnly):
    permission_classes = [IsOwnerOrReadOnly]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer

    def get_queryset(self):
        owner = self.request.user
        if owner.is_anonymous:
            return None
        return self.queryset.filter(company=owner.company_name)

class DeleteItemView(RetrieveDestroyAPIView, IsOwnerOrReadOnly):
    permission_classes = [IsOwnerOrReadOnly]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer

    def get_queryset(self):
        owner = self.request.user
        if owner.is_anonymous:
            return None
        return self.queryset.filter(company=owner.company_name)

@api_view(['GET', 'POST'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def get_excel(request):

    # Check if user is authenticated.
    if request.user.is_anonymous:
        message = {'message': 'You are not allowed to visit this page'}
        return JsonResponse(message, safe=False)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="test.xlsx"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Sample Form')
    row_num = 0
    font_style = xlwt.easyxf("font: bold on; align: wrap on, horiz left; borders: left thin, right thin, top thin, bottom thin;")
    
    columns = ['ავტორი','მანიფესტის კოდი', 'გამ. სახელი', 'გამ. გვარი', 'გამომ. ქვეყანა',
            'გამომ. ქალაქი', 'მიმღ. სახელი', 'მიმღ. გვარი', 'მიმღების ID', 'მიმღ. ქვეყანა', 'მიმღ. ქალაქი',
            'მიმღ. მისამართი','ტელ. ნომერი', 'ფასი', 'ვალუტა', 'წონა', 'კომპანია', 'ჩამოსულია', 'აღწერა']

    for col_num in range(len(columns)):
        ws.col(col_num).width = int(16*260)
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()

    wb.save(response)

    return response
'''
@api_view(['GET', 'POST'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def GeneratePdf(request, pk):
    # Check if user is authenticated.
    if request.user.is_anonymous:
        message = {'message': 'You are not allowed to visit this page'}
        return JsonResponse(message, safe=False)

    obj = Item.objects.get(id=pk)
    data = {
        'today': datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
        'customer': obj,
        'weight': obj.weight/100
    }

    template_path = 'pdf/barcode_image.html'
    response = HttpResponse(content_type='application/pdf')
    html = render_to_string(template_path, data)

    pisaStatus = pisa.CreatePDF(html, dest=response)

    return response

'''
@api_view(['GET', 'POST'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def upload_excel(request):
    if request.method == 'POST':
        #item_resource = ItemResource()
        #dataset = Dataset()
        new_items = request.FILES['myfile']

        wb = load_workbook(new_items)
        sheet_ranges = wb['Sample Form']
        print(sheet_ranges.max_row)
        number_of_items= 0
        item_id = []
        for item in sheet_ranges.iter_rows(min_row=2, values_only=True):
            value = Item.objects.create(
                owner = request.user,
                sender_name = item[0],
                sender_surname = item[1],
                company = request.user.company_name,
                sender_country = item[2],
                sender_city = item[3],
                receiver_name = item[4],
                receiver_surname = item[5],
                receiver_id = item[6],
                receiver_country = item[7],
                receiver_city = item[8],
                receiver_address = item[9],
                receiver_number = item[10],
                price = item[11],
                currency = item[12],
                weight = item[13], 
            )
            number_of_items += 1
            item_id.append(value.id)
            value.save()
    message = {'message': f'ატვირთულია {number_of_items} ამანათი', "item_id": item_id}
    return JsonResponse(message, safe=False)

@xframe_options_exempt
@api_view(['GET', 'POST'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def Generate_sticker(request, pk):
    '''
        This function takes ID of item as an argument and generates invoice
        which is a conversion of zpl command to a PDF file.

        It also utilizes third party API to make this conversion possible + a request library
    '''
    # Check if user is authenticated.
    if request.user.is_anonymous:
        message = {'message': 'You are not allowed to visit this page'}
        return JsonResponse(message, safe=False)

    # get an object with id
    obj = Item.objects.get(id=pk)

    # take an object values do generate variables in pdf
    data = [str(obj.barcode), str(obj.get_sender_full_name()),
            str(obj.sender_city), str(obj.weight),
            str(obj.get_receiver_full_name()), str(obj.receiver_number),
            str(obj.receiver_city), obj.created_at.strftime(
                "%d-%m-%Y %H:%M:%S"),
            str(obj.owner.company_name)
            ]

    # insert variables in zpl command (check utils.py to find funciton bellow)
    zpl = generate_zpl(*data)
    url = 'http://api.labelary.com/v1/printers/8dpmm/labels/4x6/0/'
    files = {'file': zpl}
    # omit this line to get PNG images back
    headers = {'Accept': 'application/pdf'}
    response = requests.post(url, headers=headers, files=files, stream=True)

    if response.status_code == 200:
        response.raw.decode_content = True
        # convert response to a readable stream
        response = HttpResponse(
            content=response, content_type=response.headers['Content-Type'])
        return response
    else:
        print('Error: ' + response.text)
        return response

@api_view(['GET', 'POST', 'PATCH'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def change_manifest(request):
    # Check if user is authenticated.
    if request.user.is_anonymous:
        message = {'message': 'You are not allowed to visit this page'}
        return JsonResponse(message, safe=False)
    elif request.method != 'PATCH':
        message = {'message': 'You can use only PATCH method'}
        return JsonResponse(message, safe=False)
    obj = Item.objects.filter(id__in=request.data['id'])
    obj.update(
        manifest_number=request.data['manifest'])
    print(request.method)
    serializer = ItemSerializer(data=obj, many=True, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['GET', 'POST', 'DELETE'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def delete_multiple_items(request):
    # Check if user is authenticated.
    if request.user.is_anonymous:
        message = {'message': 'You are not allowed to visit this page'}
        return JsonResponse(message, safe=False)
    elif request.method != 'DELETE':
        message = {'message': 'You can use only DELETE method'}
        return JsonResponse(message, safe=False)
    obj = Item.objects.filter(id__in=request.data['id'])
    obj.delete()
    serializer = ItemSerializer(data=obj, many=True, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.data, status=status.HTTP_200_OK)

def decode_base64_file(data):

    def get_file_extension(file_name, decoded_file):
        import imghdr

        extension = imghdr.what(file_name, decoded_file)
        extension = "jpg" if extension == "jpeg" else extension

        return extension

    from django.core.files.base import ContentFile
    import base64
    import six
    import uuid

    # Check if this is a base64 string
    if isinstance(data, six.string_types):
        # Check if the base64 string is in the "data:" format
        if 'data:' in data and ';base64,' in data:
            # Break out the header from the base64 content
            header, data = data.split(';base64,')

        # Try to decode the file. Return validation error if it fails.
        try:
            decoded_file = base64.b64decode(data)
        except TypeError:
            TypeError('invalid_image')

        # Generate file name:
        file_name = str(uuid.uuid4())[:12] # 12 characters are more than enough.
        # Get the file name extension:
        file_extension = get_file_extension(file_name, decoded_file)

        complete_file_name = "%s.%s" % (file_name, file_extension, )

        return ContentFile(decoded_file, name=complete_file_name)

@api_view(['GET', 'POST', 'PATCH'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def sign_document(request):
    # Check if user is authenticated.
    if request.user.is_anonymous:
        message = {'message': 'You are not allowed to visit this page'}
        return JsonResponse(message, safe=False)
    elif request.method != 'PATCH':
        message = {'message': 'You can use only PATCH method'}
        return JsonResponse(message, safe=False)
    obj = Item.objects.filter(id=request.data['id'])
    obj1 = Item.objects.get(id=request.data['id'])
    image_base64 = request.data['signature']
    obj1.signature = decode_base64_file(image_base64)
    obj1.save()
    obj.update(arrived=True, delivered=True)
    serializer = ItemSerializer(data=obj, many=True, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET', 'POST'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def export_excel(request):
    # Check if user is authenticated.
    if request.user.is_anonymous:
        message = {'message': 'You are not allowed to visit this page'}
        return JsonResponse(message, safe=False)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="test.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('ამანათები')
    row_num = 0
    font_style = xlwt.easyxf("font: bold on; align: wrap on, horiz left; borders: left thin, right thin, top thin, bottom thin;")
    items_style = xlwt.easyxf("align: wrap on, horiz left; borders: left thin, right thin, top thin, bottom thin;")

    columns = ['ID', 'ბარკოდი', 'მანიფესტის კოდი', 'გამ. სახელი', 'გამ. გვარი',
               'მიმღ. სახელი', 'მიმღ. გვარი', 'მიმღ. ქალაქი', 'მიმღ. მისამართი', 
               'მიმღების ID', 'ტელ. ნომერი', 'ფასი','ვალუტა', 'წონა', 'ავტორი', 
               'კომპანია', 'ჩამოსულია', 'აღწერა']

    for col_num in range(len(columns)):
        ws.col(col_num).width = int(16*260)
        ws.write(row_num, col_num, columns[col_num], font_style)

    font_style = xlwt.XFStyle()


    if len(request.data['id']) > 0 and request.user.is_superuser:
        rows = Item.objects.filter(id__in=request.data['id']).values_list('id', 'barcode', 'manifest_number__manifest_code', 'sender_name',
                                                                          'sender_surname', 'receiver_name', 'receiver_surname',
                                                                          'receiver_city', 'receiver_address', 'receiver_id', 'receiver_number', 'price', 'currency',
                                                                          'weight', 'owner__username', 'owner__company_name', 'arrived',
                                                                          'description')
    elif len(request.data['id']) > 0 and request.user.groups.filter(name='Company').exists():
        rows = Item.objects.filter(id__in=request.data['id']).values_list('id', 'barcode', 'manifest_number__manifest_code', 'sender_name',
                                                                          'sender_surname', 'receiver_name', 'receiver_surname',
                                                                          'receiver_city', 'receiver_address', 'receiver_id', 'receiver_number', 'price', 'currency',
                                                                          'weight', 'owner__username', 'owner__company_name',
                                                                          'arrived', 'description')

    elif len(request.data['id']) <= 0 and request.user.is_superuser and request.user.groups.filter(name='Company').exists():
        rows = Item.objects.all().values_list('id', 'barcode', 'manifest_number__manifest_code', 'sender_name',
                                              'sender_surname', 'receiver_name', 'receiver_surname',
                                              'receiver_city', 'receiver_address', 'receiver_id', 'receiver_number','price', 'currency',
                                              'weight', 'owner__username', 'owner__company_name',
                                              'arrived', 'description')

    elif len(request.data['id']) <= 0 and request.user.groups.filter(name='Company').exists():
        rows = Item.objects.filter(owner=request.user).values_list('id', 'barcode', 'manifest_number__manifest_code', 'sender_name',
                                                                          'sender_surname', 'receiver_name', 'receiver_surname',
                                                                          'receiver_city', 'receiver_address', 'receiver_id', 'receiver_number', 'price', 'currency',
                                                                          'weight', 'owner__username', 'owner__company_name',
                                                                          'arrived', 'description')
    
    else:
        rows = Item.objects.filter(owner=request.user).values_list('id', 'barcode', 'manifest_number__manifest_code', 'sender_name',
                                                                   'sender_surname', 'receiver_name', 'receiver_surname',
                                                                   'receiver_city', 'receiver_address', 'receiver_id', 'receiver_number', 'price', 'currency',
                                                                   'weight', 'owner__username', 'owner__company_name',
                                                                   'arrived', 'description')

    ws.col(2).width = int(18*260)
    for i in range(9, 16):
        ws.col(i).width = int(13*260)

    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), items_style)

    wb.save(response)

    return response

@api_view(['GET', 'POST'])
@permission_classes((IsAuthenticated, IsOwnerOrReadOnly))
def export_excel_manifest(request):
    # Check if user is authenticated.
    if request.user.is_anonymous:
        message = {'message': 'You are not allowed to visit this page'}
        return JsonResponse(message, safe=False)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="test.xlsx"'

    #create a workbook and set the style of data input
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('მანიფესტი',)
    row_num = 0
    document_cols = 0
    style = xlwt.easyxf("font: bold on; align: wrap on, horiz center; borders: left thin, right thin, top thin, bottom thin;")
    items_style = xlwt.easyxf("align: wrap on, horiz left; borders: left thin, right thin, top thin, bottom thin;")

    #columns for manifest
    manifest_columns = ['COMPANY:', 'DATE:', 'SENDER CITY:', 'RECEIVER CITY:', 
                        'MANIFEST NUMBER:', 'CMR:', 'CAR REGISTRATION No:', 'DRIVER NAME:',
                        'DRIVER SURNAME:']

    #columns for items
    columns = ['ID', 'ბარკოდი', 'მანიფესტის კოდი', 'გამ. სახელი', 'გამ. გვარი',
               'მიმღ. სახელი', 'მიმღ. გვარი', 'მიმღ. ქალაქი', 'მიმღ. მისამართი', 'მიმღების ID', 'ტელ. ნომერი', 'ფასი',
               'ვალუტა', 'წონა', 'ავტორი', 'კომპანია', 'ჩამოსულია', 'აღწერა']

    #manifest rows
    rows_manifest = Manifest.objects.filter(id=request.data['manifest_id']).values_list('owner__company_name', 'created_at',
    'sender_city', 'receiver_city', 'manifest_code', 'cmr', 'car_number', 'driver_name', 'driver_surname')
    

    #add the value of manifest in excel
    for col_num in range(len(manifest_columns)):
        ws.col(col_num).width = int(16*260)
        ws.write_merge(row_num, row_num, document_cols, 1, manifest_columns[col_num], style)
        for row in rows_manifest:
            ws.write_merge(row_num, row_num, 2, 3, str(row[col_num]), style)
            row_num += 1

    ws.write_merge(row_num, row_num, document_cols, 1, 'TOTAL ITEMS', style)
    ws.write_merge(row_num, row_num, 2, 3, str(request.data['total_items']), style)
    row_num +=1
    ws.write_merge(row_num, row_num, document_cols, 1, 'TOTAL WEIGHT', style)
    ws.write_merge(row_num, row_num, 2, 3, str(request.data['total_weight']) + ' KG', style)


    row_num +=1
    #set the remaining col width
    ws.col(2).width = int(18*260)
    ws.col(8).width = int(16*260)

    row_num +=1
    #add item columns
    for col_num in range(len(columns)):
        ws.write_merge(row_num, row_num+1, col_num, col_num, columns[col_num], style)

    row_num +=1
    #query item values that are filtered by manifest id
    rows = Item.objects.filter(manifest_number=request.data['manifest_id']).values_list('id', 'barcode', 'manifest_number__manifest_code', 'sender_name',
                                                                        'sender_surname', 'receiver_name', 'receiver_surname',
                                                                        'receiver_city', 'receiver_address', 'receiver_id', 'receiver_number', 'price', 'currency',
                                                                        'weight', 'owner__username', 'owner__company_name', 'arrived',
                                                                        'description')
    
    #add item row values
    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), items_style)

    wb.save(response)

    return response

class ManifestListView(ListAPIView, IsOwnerOrReadOnly):
    permission_classes = [IsOwnerOrReadOnly, IsAuthenticated]
    queryset = Manifest.objects.annotate(total_items=Count('item'))
    serializer_class = ManifestSerializer
    pagination_class = None

    def get_queryset(self):
        owner = self.request.user
        if owner.is_anonymous:
            return None
        return self.queryset.filter(company=owner.company_name)


class ManifestSearchView(ListAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Manifest.objects.all()
    serializer_class = ManifestSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['^manifest_code', '^owner__username',
                     '^car_number', '^sender_city', '^receiver_city']

    def get_queryset(self):
        owner = self.request.user
        if owner.is_anonymous:
            return None
        return self.queryset.filter(company=owner.company_name)


class ManifestDetailView(IsOwnerOrReadOnly, RetrieveAPIView):
    permission_classes = [IsAuthenticated, IsOwnerOrReadOnly]
    queryset = Manifest.objects.all()
    serializer_class = ManifestSerializer


class AddManifestView(CreateAPIView, IsOwnerOrReadOnly):
    permission_classes = [IsAuthenticated]
    queryset = Manifest.objects.all()
    serializer_class = ManifestSerializer

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user, company=self.request.user.company_name)


class UpdateManifestView(RetrieveUpdateAPIView, IsOwnerOrReadOnly):
    permission_classes = [IsOwnerOrReadOnly]
    queryset = Manifest.objects.all()
    serializer_class = ManifestSerializer

    def get_queryset(self):
        owner = self.request.user
        if owner.is_anonymous:
            return None
        return self.queryset.filter(company=owner.company_name)


class DeleteManifestView(RetrieveDestroyAPIView, IsOwnerOrReadOnly):
    permission_classes = [IsOwnerOrReadOnly]
    queryset = Manifest.objects.all()
    serializer_class = ManifestSerializer

    def get_queryset(self):
        owner = self.request.user
        if owner.is_anonymous:
            return None
        return self.queryset.filter(company=owner.company_name)
